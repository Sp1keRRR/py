import requests
from bs4 import BeautifulSoup
import csv
import openpyxl


def get_html(url):
    r = requests.get(url)
    return r.text


def get_total_pages(html):
    soup = BeautifulSoup(html, 'lxml')
    pagesall = soup.find('div', class_='pagebar')
    pagelast = pagesall.find_all('a', class_='page')[-1].get('href')
    total_pages = pagelast.split('=')[1]    
    return int(total_pages)


#Не работает
##def write_exl(db):
##    columns = ['title', 'price', 'city', 'data', 'url']
##    FILE_NAME = 'farpost.xlsx'
##    try:
##        wb = openpyxl.load_workbook(FILE_NAME)
##    except:
##        wb = openpyxl.Workbook()
##        # Удаление листа, создаваемого по умолчанию, при создании документа
##        for sheet_name in wb.sheetnames:
##            del wb[sheet_name]
##    # Создание нового листа
##    try:
##        ws = wb['ads']
##    except:
##        ws = wb.create_sheet('ads')
##        for i in range(1, 6):
##            ws.cell(row=1, column=i).value = columns[i-1]
##    for i in range(ws.max_row,len(db)):
##        ws.cell(row=i, column=1).value = db['title']    
##        ws.cell(row=i, column=2).value = db['price']
##        ws.cell(row=i, column=3).value = db['city']
##        ws.cell(row=i, column=4).value = db['data']
##        ws.cell(row=i, column=5).value = db['url']     
##    from openpyxl.writer.excel import save_workbook
##    save_workbook(wb, FILE_NAME)



def write_csv(db):        
    with open('farpost.txt', 'a') as file:
        writer = csv.writer(file)
        writer.writerow((db['title'],
                         db['price'],
                         db['city'],
                         db['data'],
                         db['url']))
        


def get_page_data(html):
    soup = BeautifulSoup(html, 'lxml')
    divs = soup.find('tbody', class_='native')
    ads = divs.find_all('tr', class_='bull-item -exact') 
    for ad in ads:
        try:
            title = ad.find('td', class_='descriptionCell').find('a', class_='bulletinLink auto-shy').contents[0]
        except:
            title = ''
        try:
            ref = ad.find('td', class_='descriptionCell').find('a', class_='bulletinLink auto-shy').get('href')
            url = "https://www.farpost.ru/" + ref
        except:
            url = ''
        try:
            priceFull = ad.find('td', class_='descriptionCell').find('span', attrs={"data-role":"price"}).contents[0].split(' ')
            firstPrice = priceFull[0].replace("\xa0",'')
            price = firstPrice[0].replace("\xa0",'')
        except:
            price = ''
        try:
            data = ad.find('td', class_='dateCell').find('div', class_='date').contents[0]
        except:
            data = ''
        try:
            city = ad.find('td', class_='dateCell').find('div', class_='city').contents[0]
        except:
            city = ''
        db = {'title':title,
                'price':price,
                'city':city,
                'data':data,
                'url':url}
        write_csv(db)
        

def main():
    url = "https://www.farpost.ru/job/vacancy/+/%CE%F5%F0%E0%ED%ED%E8%EA/"
    base_url = "https://www.farpost.ru/job/vacancy/+/%CE%F5%F0%E0%ED%ED%E8%EA/?page="


    total_pages = get_total_pages(get_html(url))
    
    for i in range(1, total_pages):
        url_gen = base_url + str(i) 
        html = get_html(url_gen)
        get_page_data(html)


if __name__ == '__main__':
    main()
